import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import re


def normalize_number(number):
    """Convierte cualquier número a su forma estándar: últimos 9 dígitos"""
    digits = re.sub(r'\D', '', str(number))
    return digits[-9:] if len(digits) >= 9 else digits


def load_excluded_numbers(file_path="excluded_numbers.txt"):
    """Lee y normaliza los números desde un archivo de texto"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return [normalize_number(line.strip()) for line in f if line.strip()]
    except FileNotFoundError:
        messagebox.showwarning("Aviso", "No se encontró 'excluded_numbers.txt'. No se excluirán números.")
        return []
    except Exception as e:
        messagebox.showerror("Error", f"Error al leer la lista de exclusión:\n{str(e)}")
        return []


def exclude_numbers(df, excluded_numbers):
    """Excluye filas cuyo número normalizado aparece en la lista"""
    df['NormalizedNumber'] = df['Number'].astype(str).apply(normalize_number)
    df = df[~df['NormalizedNumber'].isin(excluded_numbers)].copy()
    df.drop(columns=['NormalizedNumber'], inplace=True)
    return df


def add_hour_to_time_column(df):
    df['Time'] = df['Time'].apply(lambda t: t if len(t.split(':')) == 3 else t + ':00')
    df['datetime'] = pd.to_datetime(df['Date'] + ' ' + df['Time'], errors='coerce')
    if df['datetime'].isna().any():
        raise ValueError("Some rows have invalid 'Date' or 'Time' values.")
    df['datetime + 1h'] = df['datetime'] + pd.Timedelta(hours=2)
    return df


def add_outbound_column(df):
    df['Inbound'] = df['Inbound'].astype(bool)
    df['outbound'] = ~df['Inbound']
    df['Inbound'] = df['Inbound'].map({True: 'Yes', False: 'No'})
    df['outbound'] = df['outbound'].map({True: 'Yes', False: 'No'})
    return df


def reorder_and_select_columns(df):
    df['Answered'] = df['Answered'].map({True: 'Yes', False: 'No'})
    columns = [
        'UserName', 'UserEmail', 'UserPhone', 'Source', 'SourceDetail',
        'Date', 'Time', 'datetime + 1h', 'Duration', 'Answered',
        'Inbound', 'outbound', 'Number', 'PhonebookName'
    ]
    df = df[columns].copy()
    df = df.rename(columns={
        'UserName': 'User Name',
        'UserEmail': 'User Email',
        'UserPhone': 'User Phone',
        'Source': 'Source',
        'SourceDetail': 'Source Detail',
        'Date': 'Date',
        'Time': 'Time',
        'datetime + 1h': 'Time + 1h',
        'Duration': 'Duration (s)',
        'Answered': 'Answered',
        'Inbound': 'Incoming Calls',
        'outbound': 'Outgoing Calls',
        'Number': 'Number',
        'PhonebookName': 'Phonebook Name'
    })
    return df


def create_summary(df, writer):
    total_calls = len(df)
    total_incoming = len(df[df['Incoming Calls'] == 'Yes'])
    total_outgoing = len(df[df['Outgoing Calls'] == 'Yes'])
    total_answered = len(df[df['Answered'] == 'Yes'])
    total_missed = total_calls - total_answered

    summary_data = {
        'Metric': ['Total Calls', 'Incoming Calls', 'Outgoing Calls', 'Answered Calls', 'Missed Calls'],
        'Count': [total_calls, total_incoming, total_outgoing, total_answered, total_missed]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)


def calculate_callback_times(df):
    callbacks = []
    missed_calls = df[(df['Incoming Calls'] == 'Yes') & (df['Answered'] == 'No')].copy()
    outgoing_calls = df[df['Outgoing Calls'] == 'Yes'].copy()

    for _, missed in missed_calls.iterrows():
        missed_number = missed['Number']
        missed_time = missed['Time + 1h']
        later_outgoing = outgoing_calls[
            (outgoing_calls['Number'] == missed_number) &
            (outgoing_calls['Time + 1h'] > missed_time)
            ].sort_values(by='Time + 1h')

        if not later_outgoing.empty:
            first_callback = later_outgoing.iloc[0]
            delta = first_callback['Time + 1h'] - missed_time
            callbacks.append({
                'Date': missed_time.date(),
                'Missed Call Time': missed_time,
                'Callback Time': first_callback['Time + 1h'],
                'Delay (minutes)': int(delta.total_seconds() // 60),
                'Number': missed_number,
                'User Name': missed['User Name'],
                'Phonebook Name': missed['Phonebook Name']
            })

    callback_df = pd.DataFrame(callbacks)
    callback_df = callback_df.sort_values(by='Missed Call Time')
    return callback_df


def process_csv(input_path, output_path):
    try:
        df_original = pd.read_csv(input_path, sep='\t')
        required_columns = ['UserName', 'UserEmail', 'UserPhone', 'Source', 'SourceDetail',
                            'Date', 'Time', 'Duration', 'Answered', 'Inbound', 'Number', 'PhonebookName']

        if not all(col in df_original.columns for col in required_columns):
            raise ValueError("The CSV file does not contain the required columns.")

        excluded_numbers = load_excluded_numbers("excluded_numbers.txt")
        df = exclude_numbers(df_original.copy(), excluded_numbers)

        df = add_hour_to_time_column(df)
        df = add_outbound_column(df)
        df = reorder_and_select_columns(df)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Calls', index=False)
            create_summary(df, writer)
            callback_df = calculate_callback_times(df)
            callback_df.to_excel(writer, sheet_name='Callbacks', index=False)

            workbook = writer.book
            sheet = workbook["Callbacks"]
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for col in sheet.iter_cols(1, sheet.max_column):
                if col[0].value == "Delay (minutes)":
                    delay_col_letter = col[0].column_letter
                    break

            if 'delay_col_letter' in locals():
                sheet.conditional_formatting.add(
                    f"{delay_col_letter}2:{delay_col_letter}{sheet.max_row}",
                    CellIsRule(operator='greaterThan', formula=['40'], fill=red_fill)
                )

        messagebox.showinfo("Success", f"The Excel file has been saved at: {output_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Error during conversion: {e}")


def select_file():
    file_path = filedialog.askopenfilename(
        title="Select CSV file",
        filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*"))
    )
    csv_path.set(file_path)


def select_save_path():
    file_path = filedialog.asksaveasfilename(
        title="Save Excel file",
        defaultextension=".xlsx",
        filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
    )
    excel_path.set(file_path)


def convert_file():
    input_path = csv_path.get()
    output_path = excel_path.get()

    if not input_path or not output_path:
        messagebox.showwarning("Warning", "Please select a CSV file and a save path.")
        return

    process_csv(input_path, output_path)


# Interfaz
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

root = ctk.CTk()
root.title("CSV to Excel Converter")
root.geometry("500x400")

csv_path = ctk.StringVar()
excel_path = ctk.StringVar()

ctk.CTkLabel(root, text="CSV File:", font=("Arial", 14)).pack(pady=10)
ctk.CTkEntry(root, textvariable=csv_path, width=400).pack(pady=5)
ctk.CTkButton(root, text="Select File", command=select_file).pack(pady=5)

ctk.CTkLabel(root, text="Save as (Excel):", font=("Arial", 14)).pack(pady=10)
ctk.CTkEntry(root, textvariable=excel_path, width=400).pack(pady=5)
ctk.CTkButton(root, text="Select Path", command=select_save_path).pack(pady=5)

ctk.CTkButton(root, text="Convert File", command=convert_file, fg_color="green").pack(pady=20)

root.mainloop()
